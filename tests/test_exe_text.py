"""Tests for ExeTextHandler — text strings from EXE."""

from __future__ import annotations

import struct
from pathlib import Path

import pytest
from openpyxl import load_workbook

from gravedigger.core.handler import Manifest
from gravedigger.handlers.exe_text import (
    ExeTextHandler,
    _build_font_stub,
    _decompress_exe,
    _patch_exit_font,
)

_GAME_ROOT = Path(__file__).resolve().parent.parent / "game"
_EXE_VARIANTS = [
    ("softdisk", "1.EXE"),
    ("retail", "MANSION.EXE"),
]


def _available_exe_paths() -> list[Path]:
    return [
        _GAME_ROOT / variant / name
        for variant, name in _EXE_VARIANTS
        if (_GAME_ROOT / variant / name).exists()
    ]


@pytest.fixture()
def handler() -> ExeTextHandler:
    return ExeTextHandler()


@pytest.fixture(params=_available_exe_paths(), ids=lambda p: p.name)
def exe_path(request: pytest.FixtureRequest) -> Path:
    path: Path = request.param
    return path


@pytest.fixture()
def unpacked_dirs(handler: ExeTextHandler, exe_path: Path, tmp_path: Path) -> tuple[Path, Path]:
    translatable = tmp_path / "translatable"
    meta = tmp_path / "meta"
    translatable.mkdir()
    meta.mkdir()
    handler.unpack(exe_path, translatable, meta)
    return translatable, meta


def _read_xlsx(path: Path) -> dict[str, str]:
    """Read strings.xlsx into {id: text} dict."""
    wb = load_workbook(path)
    ws = wb.active
    assert ws is not None
    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        str_id, text = row[0], row[1]
        result[str(str_id)] = str(text) if text is not None else ""
    return result


def _write_xlsx(path: Path, strings: dict[str, str]) -> None:
    """Write {id: text} dict back to strings.xlsx."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "strings"
    ws.append(["id", "text"])
    for str_id, text in strings.items():
        ws.append([str_id, text])
    wb.save(path)


class TestUnpack:
    def test_produces_xlsx(self, unpacked_dirs: tuple[Path, Path]) -> None:
        translatable, _ = unpacked_dirs
        assert (translatable / "strings.xlsx").exists()

    def test_produces_manifest(self, unpacked_dirs: tuple[Path, Path]) -> None:
        _, meta = unpacked_dirs
        manifest = Manifest.from_json(meta / "manifest.json")
        assert manifest.handler == "ExeTextHandler"

    def test_xlsx_contains_dangerous_dave(self, unpacked_dirs: tuple[Path, Path]) -> None:
        translatable, _ = unpacked_dirs
        strings = _read_xlsx(translatable / "strings.xlsx")
        assert any("Dangerous Dave Commands" in t for t in strings.values())

    def test_xlsx_contains_game_over(self, unpacked_dirs: tuple[Path, Path]) -> None:
        translatable, _ = unpacked_dirs
        strings = _read_xlsx(translatable / "strings.xlsx")
        assert any("G A M E   O V E R" in t for t in strings.values())

    def test_xlsx_contains_copyright(self, unpacked_dirs: tuple[Path, Path]) -> None:
        translatable, _ = unpacked_dirs
        strings = _read_xlsx(translatable / "strings.xlsx")
        assert any("Softdisk" in t for t in strings.values())

    def test_xlsx_contains_level_names(self, unpacked_dirs: tuple[Path, Path]) -> None:
        translatable, _ = unpacked_dirs
        strings = _read_xlsx(translatable / "strings.xlsx")
        assert any("LEVEL 1" in t for t in strings.values())
        assert any("LEVEL 8" in t for t in strings.values())

    def test_xlsx_contains_congratulations(self, unpacked_dirs: tuple[Path, Path]) -> None:
        translatable, _ = unpacked_dirs
        strings = _read_xlsx(translatable / "strings.xlsx")
        assert any("You have freed Delbert" in t for t in strings.values())

    def test_xlsx_has_header_row(self, unpacked_dirs: tuple[Path, Path]) -> None:
        translatable, _ = unpacked_dirs
        wb = load_workbook(translatable / "strings.xlsx")
        ws = wb.active
        assert ws is not None
        assert ws.cell(1, 1).value == "id"
        assert ws.cell(1, 2).value == "text"

    def test_xlsx_string_count(self, unpacked_dirs: tuple[Path, Path]) -> None:
        translatable, _ = unpacked_dirs
        strings = _read_xlsx(translatable / "strings.xlsx")
        assert len(strings) >= 50

    def test_manifest_has_original_exe(self, unpacked_dirs: tuple[Path, Path]) -> None:
        _, meta = unpacked_dirs
        manifest = Manifest.from_json(meta / "manifest.json")
        assert "original_exe" in manifest.metadata

    def test_manifest_has_strings_meta(self, unpacked_dirs: tuple[Path, Path]) -> None:
        _, meta = unpacked_dirs
        manifest = Manifest.from_json(meta / "manifest.json")
        strings_meta = manifest.metadata["strings"]
        assert len(strings_meta) >= 50
        for entry in strings_meta:
            assert "id" in entry
            assert "offset" in entry
            assert "max_length" in entry
            assert isinstance(entry["offset"], int)
            assert isinstance(entry["max_length"], int)


class TestRepack:
    def test_roundtrip_byte_exact(
        self, handler: ExeTextHandler, exe_path: Path, tmp_path: Path
    ) -> None:
        translatable = tmp_path / "translatable"
        meta = tmp_path / "meta"
        translatable.mkdir()
        meta.mkdir()
        manifest = handler.unpack(exe_path, translatable, meta)

        repack_path = tmp_path / "repacked.exe"
        handler.repack(manifest, translatable, meta, repack_path)

        assert repack_path.read_bytes() == exe_path.read_bytes()

    def test_string_longer_than_original_relocates(
        self, handler: ExeTextHandler, exe_path: Path, tmp_path: Path
    ) -> None:
        """Strings longer than original slot are relocated to an appended block."""
        translatable = tmp_path / "translatable"
        meta = tmp_path / "meta"
        translatable.mkdir()
        meta.mkdir()
        manifest = handler.unpack(exe_path, translatable, meta)

        xlsx_path = translatable / "strings.xlsx"
        strings = _read_xlsx(xlsx_path)
        long_text = "A very long replacement string that exceeds the original"
        strings["win_line13"] = long_text
        _write_xlsx(xlsx_path, strings)

        repack_path = tmp_path / "repacked.exe"
        handler.repack(manifest, translatable, meta, repack_path)

        # Output is an uncompressed EXE when relocation is needed
        repacked = repack_path.read_bytes()
        assert long_text.encode("ascii") in repacked

    def test_patching_decompressed_strings(
        self, handler: ExeTextHandler, exe_path: Path, tmp_path: Path
    ) -> None:
        import struct

        exe_data = exe_path.read_bytes()
        decompressed = _decompress_exe(exe_data)

        header_para = struct.unpack_from("<H", decompressed, 8)[0]
        code_start = header_para * 16

        patched = bytearray(decompressed)

        win_offset = code_start + 0x27646
        assert patched[win_offset : win_offset + 8] == b"You win!"
        patched[win_offset : win_offset + 8] = b"U win!!!"
        assert patched[win_offset : win_offset + 9] == b"U win!!!\x00"

        god_offset = code_start + 0x273AE
        assert patched[god_offset : god_offset + 12] == b"God mode off"
        patched[god_offset : god_offset + 13] = b"Off" + b"\x00" * 10
        assert patched[god_offset : god_offset + 13] == b"Off" + b"\x00" * 10

        assert b"U win!!!" in patched
        assert b"You win!" not in patched
        assert b"Off" in patched[god_offset : god_offset + 13]

    def test_modify_string_and_repack(
        self, handler: ExeTextHandler, exe_path: Path, tmp_path: Path
    ) -> None:
        translatable = tmp_path / "translatable"
        meta = tmp_path / "meta"
        translatable.mkdir()
        meta.mkdir()
        manifest = handler.unpack(exe_path, translatable, meta)

        xlsx_path = translatable / "strings.xlsx"
        strings = _read_xlsx(xlsx_path)
        strings["copyright"] = strings["copyright"].replace("Softdisk", "Xoftdisk")
        _write_xlsx(xlsx_path, strings)

        repack_path = tmp_path / "repacked.exe"
        handler.repack(manifest, translatable, meta, repack_path)

        repacked_dec = _decompress_exe(repack_path.read_bytes())
        assert b"Xoftdisk" in repacked_dec
        assert b"Softdisk" not in repacked_dec

    def test_repack_preserves_other_data(
        self, handler: ExeTextHandler, exe_path: Path, tmp_path: Path
    ) -> None:
        import struct

        translatable = tmp_path / "translatable"
        meta = tmp_path / "meta"
        translatable.mkdir()
        meta.mkdir()
        manifest = handler.unpack(exe_path, translatable, meta)

        repack_path = tmp_path / "repacked.exe"
        handler.repack(manifest, translatable, meta, repack_path)

        original_dec = _decompress_exe(exe_path.read_bytes())
        repacked_dec = _decompress_exe(repack_path.read_bytes())

        header_para = struct.unpack_from("<H", original_dec, 8)[0]
        code_start = header_para * 16
        death_offset = code_start + 0x18A60
        block_size = 0x33B0 * 4

        assert (
            original_dec[death_offset : death_offset + block_size]
            == repacked_dec[death_offset : death_offset + block_size]
        )


class TestFontStub:
    def test_returns_bytes(self) -> None:
        result = _build_font_stub(0x1234)
        assert isinstance(result, bytes)

    def test_starts_with_mode_set(self) -> None:
        result = _build_font_stub(0x1234)
        assert result[:5] == b"\xb8\x03\x00\xcd\x10"

    def test_ends_with_retf(self) -> None:
        result = _build_font_stub(0x1234)
        assert result[-1:] == b"\xcb"

    def test_contains_one_int10h(self) -> None:
        result = _build_font_stub(0x1234)
        count = 0
        for i in range(len(result) - 1):
            if result[i] == 0xCD and result[i + 1] == 0x10:
                count += 1
        assert count == 1

    def test_font_seg_embedded(self) -> None:
        result = _build_font_stub(0x1234)
        idx = result.index(b"\xb8\x34\x12")
        assert idx > 4

    def test_font_seg_different_value(self) -> None:
        result = _build_font_stub(0xABCD)
        assert b"\xb8\xcd\xab" in result

    def test_total_length(self) -> None:
        result = _build_font_stub(0x1234)
        assert len(result) == 165


class TestPatchExitFont:
    """Tests for _patch_exit_font on real game EXEs."""

    @staticmethod
    def _decompress(exe_path: Path) -> tuple[bytearray, int]:
        data = exe_path.read_bytes()
        dec = bytearray(_decompress_exe(data))
        code_start = struct.unpack_from("<H", dec, 8)[0] * 16
        return dec, code_start

    def test_far_call_written(self, exe_path: Path) -> None:
        exe, code_start = self._decompress(exe_path)
        _patch_exit_font(exe, code_start, bytes(4096))
        assert exe[code_start + 0x679] == 0x9A

    def test_far_call_offset_is_zero(self, exe_path: Path) -> None:
        exe, code_start = self._decompress(exe_path)
        _patch_exit_font(exe, code_start, bytes(4096))
        call_addr = code_start + 0x679
        offset = struct.unpack_from("<H", exe, call_addr + 1)[0]
        assert offset == 0x0000

    def test_far_call_segment_points_to_stub(self, exe_path: Path) -> None:
        exe, code_start = self._decompress(exe_path)
        original_len = len(exe)
        font_data = bytes(4096)
        _patch_exit_font(exe, code_start, font_data)

        append_offset = (original_len + 15) & ~15
        stub_file_offset = append_offset + len(font_data)
        expected_seg = (stub_file_offset - code_start) // 16

        call_addr = code_start + 0x679
        actual_seg = struct.unpack_from("<H", exe, call_addr + 3)[0]
        assert actual_seg == expected_seg

    def test_font_appended_paragraph_aligned(self, exe_path: Path) -> None:
        exe, code_start = self._decompress(exe_path)
        original_len = len(exe)
        font_data = bytes(range(256)) * 16  # 4096 bytes, recognizable pattern
        _patch_exit_font(exe, code_start, font_data)

        assert len(exe) > original_len
        append_start = (original_len + 15) & ~15
        assert append_start % 16 == 0
        assert exe[append_start : append_start + 4096] == font_data

    def test_stub_follows_font(self, exe_path: Path) -> None:
        exe, code_start = self._decompress(exe_path)
        original_len = len(exe)
        _patch_exit_font(exe, code_start, bytes(4096))

        append_start = (original_len + 15) & ~15
        stub_start = append_start + 4096
        assert exe[stub_start : stub_start + 5] == b"\xb8\x03\x00\xcd\x10"
        stub = _build_font_stub(0)
        assert exe[stub_start + len(stub) - 1] == 0xCB

    def test_two_reloc_entries_added(self, exe_path: Path) -> None:
        exe, code_start = self._decompress(exe_path)
        orig_crlc = struct.unpack_from("<H", exe, 6)[0]
        _patch_exit_font(exe, code_start, bytes(4096))
        assert struct.unpack_from("<H", exe, 6)[0] == orig_crlc + 2

    def test_reloc_entry1_targets_far_call_segment(self, exe_path: Path) -> None:
        exe, code_start = self._decompress(exe_path)
        orig_crlc = struct.unpack_from("<H", exe, 6)[0]
        _patch_exit_font(exe, code_start, bytes(4096))
        e_lfarlc = struct.unpack_from("<H", exe, 0x18)[0]
        r1_off, r1_seg = struct.unpack_from("<HH", exe, e_lfarlc + orig_crlc * 4)
        assert r1_seg * 16 + r1_off == 0x67C

    def test_reloc_entry2_targets_font_seg_in_stub(self, exe_path: Path) -> None:
        exe, code_start = self._decompress(exe_path)
        original_len = len(exe)
        orig_crlc = struct.unpack_from("<H", exe, 6)[0]
        _patch_exit_font(exe, code_start, bytes(4096))

        e_lfarlc = struct.unpack_from("<H", exe, 0x18)[0]
        r2_off, r2_seg = struct.unpack_from("<HH", exe, e_lfarlc + (orig_crlc + 1) * 4)
        append_offset = (original_len + 15) & ~15
        stub_code_offset = (append_offset + 4096) - code_start
        from gravedigger.handlers.exe_text import _FONT_STUB_SEG_OFFSET

        expected = stub_code_offset + _FONT_STUB_SEG_OFFSET
        assert r2_seg * 16 + r2_off == expected

    def test_file_size_in_header_updated(self, exe_path: Path) -> None:
        exe, code_start = self._decompress(exe_path)
        _patch_exit_font(exe, code_start, bytes(4096))
        file_size = len(exe)
        assert struct.unpack_from("<H", exe, 2)[0] == file_size & 0x1FF
        assert struct.unpack_from("<H", exe, 4)[0] == (file_size + 0x1FF) >> 9

    def test_font_seg_in_stub_matches_font_position(self, exe_path: Path) -> None:
        exe, code_start = self._decompress(exe_path)
        original_len = len(exe)
        _patch_exit_font(exe, code_start, bytes(4096))

        append_offset = (original_len + 15) & ~15
        expected_font_seg = (append_offset - code_start) // 16

        from gravedigger.handlers.exe_text import _FONT_STUB_SEG_OFFSET

        stub_start = append_offset + 4096
        actual_font_seg = struct.unpack_from("<H", exe, stub_start + _FONT_STUB_SEG_OFFSET)[0]
        assert actual_font_seg == expected_font_seg

    def test_resize_constant_updated(self, exe_path: Path) -> None:
        """Data-size constant in startup code covers appended font+stub."""
        from gravedigger.handlers.exe_text import (
            _DS_BASE,
            _RESIZE_IMM_OFFSET,
        )

        exe, code_start = self._decompress(exe_path)
        original_len = len(exe)
        _patch_exit_font(exe, code_start, bytes(4096))

        append_offset = (original_len + 15) & ~15
        stub = _build_font_stub(0)
        stub_end_code = (append_offset + 4096 + len(stub)) - code_start
        data_size = struct.unpack_from("<H", exe, code_start + _RESIZE_IMM_OFFSET)[0]
        assert data_size >= stub_end_code - _DS_BASE

    def test_startup_jne_patched_to_jmp(self, exe_path: Path) -> None:
        """JNE at 0x8C is patched to JMP to prevent DI override / SP=0 wrap."""
        from gravedigger.handlers.exe_text import _RESIZE_JNE_OFFSET

        exe, code_start = self._decompress(exe_path)
        assert exe[code_start + _RESIZE_JNE_OFFSET] == 0x75  # JNE before
        _patch_exit_font(exe, code_start, bytes(4096))
        assert exe[code_start + _RESIZE_JNE_OFFSET] == 0xEB  # JMP after


class TestFilePatterns:
    def test_file_patterns(self, handler: ExeTextHandler) -> None:
        assert "DAVE.EXE" in handler.file_patterns
        assert "1.EXE" in handler.file_patterns
        assert "MANSION.EXE" in handler.file_patterns


_EXIT_SCREEN_OFFSET = 0x112F0
_EXIT_SCREEN_SIZE = 4000  # 80 * 25 * 2
_EXIT_SCREEN_WIDTH = 80
_EXIT_SCREEN_HEIGHT = 25
_EXIT_SCREEN_FONT_HEIGHT = 16


class TestExitScreen:
    """Tests for exit screen unpack/repack in ExeTextHandler."""

    def test_unpack_produces_exit_screen_xb(self, unpacked_dirs: tuple[Path, Path]) -> None:
        translatable, _ = unpacked_dirs
        assert (translatable / "exit_screen.xb").exists()

    def test_exit_screen_xb_is_valid_xbin(self, unpacked_dirs: tuple[Path, Path]) -> None:
        from gravedigger.xbin import parse

        translatable, _ = unpacked_dirs
        data = (translatable / "exit_screen.xb").read_bytes()
        result = parse(data)
        assert result is not None

    def test_exit_screen_dimensions(self, unpacked_dirs: tuple[Path, Path]) -> None:
        from gravedigger.xbin import parse

        translatable, _ = unpacked_dirs
        data = (translatable / "exit_screen.xb").read_bytes()
        result = parse(data)
        assert result.width == _EXIT_SCREEN_WIDTH
        assert result.height == _EXIT_SCREEN_HEIGHT

    def test_exit_screen_has_font(self, unpacked_dirs: tuple[Path, Path]) -> None:
        from gravedigger.xbin import parse

        translatable, _ = unpacked_dirs
        data = (translatable / "exit_screen.xb").read_bytes()
        result = parse(data)
        assert result.font is not None
        assert result.font_height == _EXIT_SCREEN_FONT_HEIGHT

    def test_exit_screen_image_data_size(self, unpacked_dirs: tuple[Path, Path]) -> None:
        from gravedigger.xbin import parse

        translatable, _ = unpacked_dirs
        data = (translatable / "exit_screen.xb").read_bytes()
        result = parse(data)
        assert len(result.image_data) == _EXIT_SCREEN_SIZE

    def test_exit_screen_roundtrip_byte_exact(
        self, handler: ExeTextHandler, exe_path: Path, tmp_path: Path
    ) -> None:
        """Repack without modifications produces byte-exact original."""
        translatable = tmp_path / "translatable"
        meta = tmp_path / "meta"
        translatable.mkdir()
        meta.mkdir()
        manifest = handler.unpack(exe_path, translatable, meta)

        repack_path = tmp_path / "repacked.exe"
        handler.repack(manifest, translatable, meta, repack_path)

        assert repack_path.read_bytes() == exe_path.read_bytes()

    def test_exit_screen_image_data_matches_exe(
        self, handler: ExeTextHandler, exe_path: Path, tmp_path: Path
    ) -> None:
        """Image data in exit_screen.xb matches raw bytes in decompressed EXE."""
        import struct as _struct

        from gravedigger.xbin import parse

        translatable = tmp_path / "translatable"
        meta = tmp_path / "meta"
        translatable.mkdir()
        meta.mkdir()
        handler.unpack(exe_path, translatable, meta)

        exe_data = exe_path.read_bytes()
        decompressed = _decompress_exe(exe_data)
        header_para = _struct.unpack_from("<H", decompressed, 8)[0]
        code_start = header_para * 16
        abs_offset = code_start + _EXIT_SCREEN_OFFSET
        raw_screen = decompressed[abs_offset : abs_offset + _EXIT_SCREEN_SIZE]

        data = (translatable / "exit_screen.xb").read_bytes()
        result = parse(data)
        assert result.image_data == raw_screen

    def test_modify_exit_screen_changes_exe(
        self, handler: ExeTextHandler, exe_path: Path, tmp_path: Path
    ) -> None:
        """Modifying exit_screen.xb changes correct bytes in the repacked EXE."""
        import struct as _struct

        from gravedigger.xbin import build, parse

        translatable = tmp_path / "translatable"
        meta = tmp_path / "meta"
        translatable.mkdir()
        meta.mkdir()
        manifest = handler.unpack(exe_path, translatable, meta)

        xb_path = translatable / "exit_screen.xb"
        orig = parse(xb_path.read_bytes())

        # Flip first byte of image data
        modified_image = bytes([orig.image_data[0] ^ 0xFF]) + orig.image_data[1:]
        new_xb = build(
            orig.width,
            orig.height,
            modified_image,
            font=orig.font,
            font_height=orig.font_height,
        )
        xb_path.write_bytes(new_xb)

        repack_path = tmp_path / "repacked.exe"
        handler.repack(manifest, translatable, meta, repack_path)

        # When exit screen is modified the output is a decompressed EXE; try to
        # decompress but fall back to reading raw if already decompressed.
        repacked_raw = repack_path.read_bytes()
        try:
            repacked_dec = bytearray(_decompress_exe(repacked_raw))
        except ValueError:
            repacked_dec = bytearray(repacked_raw)

        orig_dec = bytearray(_decompress_exe(exe_path.read_bytes()))
        header_para = _struct.unpack_from("<H", bytes(orig_dec), 8)[0]
        code_start = header_para * 16
        abs_offset = code_start + _EXIT_SCREEN_OFFSET

        assert repacked_dec[abs_offset] == orig.image_data[0] ^ 0xFF
        assert repacked_dec[abs_offset + 1 : abs_offset + _EXIT_SCREEN_SIZE] == orig.image_data[1:]


class TestFontIntegration:
    """Integration tests for font embedding during repack."""

    def test_repack_with_font_embeds_stub(
        self, handler: ExeTextHandler, exe_path: Path, tmp_path: Path
    ) -> None:
        """Repack with XBIN containing a font embeds font+stub in output EXE."""
        import struct as _struct

        from gravedigger.xbin import build, parse

        translatable = tmp_path / "translatable"
        meta = tmp_path / "meta"
        translatable.mkdir()
        meta.mkdir()
        manifest = handler.unpack(exe_path, translatable, meta)

        # Modify exit screen image to force needs_relocation path
        xb_path = translatable / "exit_screen.xb"
        orig = parse(xb_path.read_bytes())
        modified_image = bytes([orig.image_data[0] ^ 0xFF]) + orig.image_data[1:]
        assert orig.font is not None
        new_xb = build(
            orig.width,
            orig.height,
            modified_image,
            font=orig.font,
            font_height=orig.font_height,
        )
        xb_path.write_bytes(new_xb)

        repack_path = tmp_path / "repacked.exe"
        handler.repack(manifest, translatable, meta, repack_path)

        repacked = repack_path.read_bytes()
        header_para = _struct.unpack_from("<H", repacked, 8)[0]
        code_start = header_para * 16

        # Far call at offset 0x679
        assert repacked[code_start + 0x679] == 0x9A

        # Font data is present in the output
        assert orig.font in repacked

        # Stub ends with RETF (0xCB)
        stub = _build_font_stub(0)
        # Find the stub by its mode-set prefix after the font
        font_pos = repacked.index(orig.font)
        stub_start = font_pos + len(orig.font)
        assert repacked[stub_start : stub_start + 5] == b"\xb8\x03\x00\xcd\x10"
        assert repacked[stub_start + len(stub) - 1] == 0xCB

        # Relocation count increased by 2
        e_crlc = _struct.unpack_from("<H", repacked, 6)[0]
        orig_dec = _decompress_exe(exe_path.read_bytes())
        orig_crlc = _struct.unpack_from("<H", orig_dec, 6)[0]
        assert e_crlc == orig_crlc + 2

    def test_repack_without_font_no_patch(
        self, handler: ExeTextHandler, exe_path: Path, tmp_path: Path
    ) -> None:
        """Repack with XBIN without font does not patch the EXE with font stub."""
        import struct as _struct

        from gravedigger.xbin import build, parse

        translatable = tmp_path / "translatable"
        meta = tmp_path / "meta"
        translatable.mkdir()
        meta.mkdir()
        manifest = handler.unpack(exe_path, translatable, meta)

        # Rebuild XBIN without font but with modified image to force needs_relocation
        xb_path = translatable / "exit_screen.xb"
        orig = parse(xb_path.read_bytes())
        modified_image = bytes([orig.image_data[0] ^ 0xFF]) + orig.image_data[1:]
        new_xb = build(
            orig.width,
            orig.height,
            modified_image,
            font=None,  # No font
            font_height=orig.font_height,
        )
        xb_path.write_bytes(new_xb)

        repack_path = tmp_path / "repacked.exe"
        handler.repack(manifest, translatable, meta, repack_path)

        repacked = repack_path.read_bytes()
        header_para = _struct.unpack_from("<H", repacked, 8)[0]
        code_start = header_para * 16

        # Original opcodes should remain (no far call)
        assert repacked[code_start + 0x679 : code_start + 0x679 + 3] == b"\xb8\x03\x00"
