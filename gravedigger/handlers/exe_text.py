"""ExeTextHandler — text strings from EXE.

Extracts and patches NUL-terminated ASCII text strings embedded in the
decompressed EXE (PKLITE or LZEXE compressed). Strings are identified by
their code image offsets, which are hardcoded based on analysis of the
Dangerous Dave executable. The code image is identical across known variants.

The string table covers: UI text, menus, prompts, game messages, win/lose
screens, copyright, configuration text, and error messages.
"""

from __future__ import annotations

import base64
import struct
from typing import TYPE_CHECKING, Any, ClassVar

from openpyxl import Workbook, load_workbook

from gravedigger.compression import lzexe, pklite
from gravedigger.core.handler import FormatHandler, Manifest
from gravedigger.data import load_cp866_font
from gravedigger.xbin import build as xbin_build
from gravedigger.xbin import parse as xbin_parse

if TYPE_CHECKING:
    from pathlib import Path


# Each entry: (code_offset, description, xrefs)
# code_offset is relative to the start of the code image in the decompressed EXE.
# xrefs are code offsets of the 16-bit immediate operand in `mov ax, imm16` (opcode B8)
# instructions that reference this string via its DS-relative offset.
# DS base segment = 0x2711 (linear 0x27110).
_DS_BASE = 0x27110

_EXIT_SCREEN_OFFSET = 0x112F0  # code offset of exit screen in decompressed EXE
_EXIT_SCREEN_WIDTH = 80
_EXIT_SCREEN_HEIGHT = 25
_EXIT_SCREEN_FONT_HEIGHT = 16
_EXIT_SCREEN_SIZE = _EXIT_SCREEN_WIDTH * _EXIT_SCREEN_HEIGHT * 2  # 4000

_STRING_TABLE: list[tuple[int, str, list[int]]] = [
    # Help / commands screen
    (0x27242, "help_title", [0x00248]),
    (0x27275, "help_f2", [0x0025A]),
    (0x2728B, "help_f3", [0x00263]),
    (0x272A9, "help_f4", [0x0026C]),
    (0x272BE, "help_f5", [0x00275]),
    (0x272D0, "help_esc", [0x0027E]),
    (0x272DC, "help_tab", [0x00287]),
    (0x272F2, "help_ctrl", [0x00290]),
    (0x27309, "help_down_ctrl", [0x00299]),
    (0x27324, "help_alt", [0x002A2]),
    (0x2733B, "help_up_down", [0x002AB]),
    (0x27350, "help_left_right", [0x002B4]),
    # Memory usage
    (0x27366, "memory_title", [0x002D8]),
    # God mode
    (0x273AE, "god_mode_off", [0x0035E]),
    (0x273BB, "god_mode_on", [0x00363]),
    # Prompts
    (0x273C7, "warp_prompt", [0x003CC]),
    (0x273E1, "sound_prompt", [0x00491]),
    (0x273EE, "reset_game_prompt", [0x004F2]),
    (0x27400, "quit_prompt", [0x00531]),
    # Status bar
    (0x2740D, "score_label", [0x0057B]),
    (0x27417, "high_score_label", [0x005AC]),
    (0x27426, "daves_label", [0x005DD]),
    # Disk I/O error
    (0x2742D, "disk_error_line1", [0x006F3]),
    (0x2744D, "disk_error_line2", [0x006FC]),
    # Tile error
    (0x27473, "tile_error", [0x008F4]),
    # Win screen
    (0x27510, "win_congrats", [0x00F28, 0x010FA]),
    (0x27522, "win_line1", [0x01103]),
    (0x2753F, "win_line2", [0x0110C]),
    (0x27554, "win_line3", [0x01115]),
    (0x27564, "win_line4", [0x01132]),
    (0x2757B, "win_line5", [0x0113B]),
    (0x27596, "win_line6", [0x01144]),
    (0x275AC, "win_line7", [0x0114D]),
    (0x275C2, "win_line8", [0x0116F]),
    (0x275DE, "win_line9", [0x01178]),
    (0x275FC, "win_line10", [0x0119A]),
    (0x27618, "win_line11", [0x011A3]),
    (0x27631, "win_line12", [0x011AC]),
    (0x27646, "win_line13", [0x011CE]),
    # Level loading
    (0x2765B, "level_1", [0x01304]),
    (0x27664, "level_2", [0x01309]),
    (0x2766D, "level_3", [0x0130E]),
    (0x27676, "level_4", [0x01313]),
    (0x2767F, "level_5", [0x01318]),
    (0x27688, "level_6", [0x0131D]),
    (0x27691, "level_7", [0x01322]),
    (0x2769A, "level_8", [0x01327]),
    (0x276A3, "loading", [0x01482]),
    # Startup messages
    (0x276C2, "startup_executing", [0x015F1]),
    (0x276E1, "startup_joy_yes", [0x01613]),
    (0x276F3, "startup_joy_no", [0x01624]),
    (0x27709, "startup_ega", [0x01640]),
    (0x2771B, "startup_vga", [0x01650]),
    (0x2772D, "startup_no_card", [0x01659]),
    (0x27778, "startup_no_card_prompt", [0x01662]),
    # Debug / ted
    (0x277B5, "ted_exit", [0x01759]),
    (0x277C8, "bad_react", [0x0177E]),
    (0x277D9, "object_overflow", [0x017BB]),
    # Game over screen
    (0x277F5, "game_over", [0x02345]),
    (0x27807, "new_high_score", [0x023BB]),
    (0x27829, "game_over_congrats", [0x02430]),
    (0x2783C, "continue_level", [0x024F9]),
    # Copyright and info
    (0x29846, "copyright", [0x067C5]),
    (0x29868, "info_hint", [0x067CE]),
    # Gamer's Edge ad
    (0x298AE, "ad_line1", [0x06B3E]),
    (0x298C7, "ad_line2", [0x06B47]),
    (0x298E1, "ad_line3", [0x06B50]),
    (0x298F9, "ad_line4", [0x06B59]),
    (0x29912, "ad_line5", [0x06B62]),
    (0x2992A, "ad_line6", [0x06B6B]),
    # Joystick configuration
    (0x29A46, "joy_config_title", [0x06BBA]),
    (0x29A76, "joy_hold_line1", [0x06BCC]),
    (0x29A90, "joy_upper_left", [0x06BD5]),
    (0x29AA7, "joy_press_button", [0x06BDE]),
    (0x29AD3, "joy_lower_right", [0x06CA3]),
    (0x29AEB, "joy_press_button2", [0x06CAC]),
    (0x29AFB, "joy_button_choice", [0x06DB3]),
    # Keyboard configuration
    (0x29BA5, "kbd_config_title", [0x06F61]),
    (0x29C0B, "kbd_modify_action", [0x06FAE]),
    (0x29C28, "kbd_press_key", [0x070AC]),
    # Error messages
    (0x29C3E, "bload_error", [0x077C3]),
    (0x29C5F, "huff_error", [0x079A4]),
    (0x29C87, "rlew_error", [0x07BB7]),
]


_LZEXE_SIG = b"LZ91"
_PKLITE_SIG = b"PKLITE"


def _is_lzexe(data: bytes) -> bool:
    return len(data) >= 0x20 and data[0x1C:0x20] == _LZEXE_SIG


def _decompress_exe(data: bytes) -> bytes:
    if _is_lzexe(data):
        return lzexe.decompress(data)
    return pklite.decompress(data)


def _compress_exe(decompressed: bytes, original: bytes) -> bytes:
    if _is_lzexe(original):
        return lzexe.compress(decompressed, original)
    return pklite.compress(decompressed, original)


def _read_nul_string(data: bytes, offset: int) -> str:
    """Read a NUL-terminated ASCII string from data at offset."""
    end = data.index(0, offset)
    return data[offset:end].decode("ascii")


class ExeTextHandler(FormatHandler):
    """Handler for text strings embedded in the game EXE."""

    file_patterns: ClassVar[list[str]] = ["DAVE.EXE", "1.EXE", "MANSION.EXE"]

    def unpack(self, input_path: Path, translatable_dir: Path, meta_dir: Path) -> Manifest:
        exe_data = input_path.read_bytes()
        decompressed = _decompress_exe(exe_data)

        header_para = struct.unpack_from("<H", decompressed, 8)[0]
        code_start = header_para * 16

        strings_meta: list[dict[str, Any]] = []
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "strings"
        ws.append(["id", "text"])

        for code_offset, name, _xrefs in _STRING_TABLE:
            abs_offset = code_start + code_offset
            text = _read_nul_string(decompressed, abs_offset)
            stripped = text.rstrip("\n")
            trailing_newlines = len(text) - len(stripped)
            ws.append([name, stripped])
            strings_meta.append(
                {
                    "id": name,
                    "offset": code_offset,
                    "max_length": len(text),
                    "trailing_newlines": trailing_newlines,
                }
            )

        wb.save(translatable_dir / "strings.xlsx")

        # Extract exit screen (80x25 VGA text-mode B800 data) as XBIN
        exit_abs = code_start + _EXIT_SCREEN_OFFSET
        image_data = decompressed[exit_abs : exit_abs + _EXIT_SCREEN_SIZE]
        font = load_cp866_font()
        xbin_bytes = xbin_build(
            _EXIT_SCREEN_WIDTH,
            _EXIT_SCREEN_HEIGHT,
            image_data,
            font=font,
            font_height=_EXIT_SCREEN_FONT_HEIGHT,
        )
        (translatable_dir / "exit_screen.xb").write_bytes(xbin_bytes)

        metadata: dict[str, Any] = {
            "original_exe": base64.b64encode(exe_data).decode(),
            "strings": strings_meta,
        }

        manifest = Manifest(
            handler="ExeTextHandler",
            source_file=input_path.name,
            metadata=metadata,
        )
        manifest.to_json(meta_dir / "manifest.json")
        return manifest

    def repack(
        self, manifest: Manifest, translatable_dir: Path, meta_dir: Path, output_path: Path
    ) -> None:
        meta = manifest.metadata
        original_exe = base64.b64decode(meta["original_exe"])
        decompressed = bytearray(_decompress_exe(original_exe))

        header_para = struct.unpack_from("<H", decompressed, 8)[0]
        code_start = header_para * 16
        code_image_size = len(decompressed) - code_start

        # Build lookup from XLSX: id -> text
        wb = load_workbook(translatable_dir / "strings.xlsx")
        ws = wb.active
        assert ws is not None
        text_by_id: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid, text = row[0], row[1]
            text_by_id[str(rid)] = str(text) if text is not None else ""

        strings_meta: list[dict[str, Any]] = meta["strings"]

        # Prepare encoded strings with trailing newlines restored
        encoded_by_id: dict[str, bytes] = {}
        for entry in strings_meta:
            str_id = entry["id"]
            new_text = text_by_id[str_id]
            trailing = entry.get("trailing_newlines", 0)
            if trailing and not new_text.endswith("\n"):
                new_text += "\n" * trailing
            encoded_by_id[str_id] = new_text.encode("ascii")

        # Check if any string exceeds its original slot
        needs_relocation = any(len(encoded_by_id[e["id"]]) > e["max_length"] for e in strings_meta)

        # Patch exit screen from exit_screen.xb
        xb_path = translatable_dir / "exit_screen.xb"
        xb = xbin_parse(xb_path.read_bytes())
        exit_abs = code_start + _EXIT_SCREEN_OFFSET
        original_screen = bytes(decompressed[exit_abs : exit_abs + _EXIT_SCREEN_SIZE])
        exit_screen_changed = xb.image_data != original_screen
        decompressed[exit_abs : exit_abs + _EXIT_SCREEN_SIZE] = xb.image_data

        # If exit screen data changed, we cannot re-compress (compressor back-references
        # may reference the modified region). Use the decompressed-EXE output path instead.
        if exit_screen_changed:
            needs_relocation = True

        if not needs_relocation:
            # All strings fit in-place — use the simple path
            for entry in strings_meta:
                str_id = entry["id"]
                code_offset: int = entry["offset"]
                max_length: int = entry["max_length"]
                encoded = encoded_by_id[str_id]
                abs_offset = code_start + code_offset
                pad_len = max_length - len(encoded)
                decompressed[abs_offset : abs_offset + max_length + 1] = encoded + b"\x00" * (
                    pad_len + 1
                )
            result = _compress_exe(bytes(decompressed), original_exe)
        else:
            # Hybrid path: patch strings that fit in-place, relocate only
            # those that exceed their original slot.  Old string content is
            # preserved (not zeroed) so that any unknown cross-references
            # still find valid data.
            extra_block = bytearray()
            relocated_ids: set[str] = set()
            new_offsets: dict[str, int] = {}

            # First pass: identify which strings need relocation and build
            # the appended block for them.
            for entry in strings_meta:
                str_id = entry["id"]
                encoded = encoded_by_id[str_id]
                max_length = entry["max_length"]
                if len(encoded) > max_length:
                    # String too long — relocate to appended block
                    new_code_offset = code_image_size + len(extra_block)
                    new_offsets[str_id] = new_code_offset - _DS_BASE
                    extra_block.extend(encoded)
                    extra_block.append(0)  # NUL terminator
                    relocated_ids.add(str_id)

            # Second pass: patch all strings.
            for entry in strings_meta:
                str_id = entry["id"]
                code_offset = entry["offset"]
                max_length = entry["max_length"]
                encoded = encoded_by_id[str_id]
                abs_offset = code_start + code_offset
                if str_id not in relocated_ids:
                    # Fits in place — overwrite with padding
                    pad_len = max_length - len(encoded)
                    decompressed[abs_offset : abs_offset + max_length + 1] = encoded + b"\x00" * (
                        pad_len + 1
                    )

            # Patch xrefs only for relocated strings.
            for _code_offset, name, xrefs in _STRING_TABLE:
                if name not in relocated_ids:
                    continue
                new_ds_off = new_offsets[name]
                if new_ds_off > 0xFFFF:  # pragma: no cover
                    msg = f"String {name!r} at DS offset 0x{new_ds_off:x} exceeds 16-bit range"
                    raise ValueError(msg)
                for xref_offset in xrefs:
                    abs_offset = code_start + xref_offset
                    struct.pack_into("<H", decompressed, abs_offset, new_ds_off)

            # Output the decompressed EXE with the extra string block appended.
            # Re-encoding into the compression format is impractical because
            # the decompressor stubs have hardcoded size values.
            decompressed.extend(extra_block)
            file_size = len(decompressed)
            struct.pack_into("<HH", decompressed, 2, file_size & 0x1FF, (file_size + 0x1FF) >> 9)
            # Fix MinAlloc/MaxAlloc: the decompressed header was only used by
            # the PKLITE/LZEXE stub internally and may have invalid values
            # (e.g. MinAlloc > MaxAlloc).  Set MaxAlloc=0xFFFF to request all
            # available memory, and MinAlloc=0 since the load module already
            # contains the full code image including the appended strings.
            struct.pack_into("<HH", decompressed, 0x0A, 0x0000, 0xFFFF)
            result = bytes(decompressed)

        output_path.write_bytes(result)
